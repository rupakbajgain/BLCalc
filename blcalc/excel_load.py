"""
Parse Excel Various Formats and save in common format
"""

from pathlib import Path

import xlrd
import openpyxl

from .helper import TempNameGenerator

class SCell:
    """
    Represent a cell of spreadsheet
    """
    def __init__(self, ty, val):
        """
        ty: type of cell
        val: field value
        """
        self.ctype = ty
        self.value = val

    def __repr__(self):
        """
        display as
        """
        return str(self.value)

    def get_value(self):
        """
        Get value of cell
        """
        return self.value

class BoreholeDataSheets:
    """
    Represent Borehole log data,
    Represents more than one one sheet
    - for one location

    properties:
    sheets: list of sheets, dict name and value
    """
    def __init__(self):
        """
        Create empty data
        """
        self.sheets={}

    @staticmethod
    def load_file_xls(filename):
        """
        Open file using xlrd for older excel format
        """
        sheet_list={}
        workbook = xlrd.open_workbook(filename, formatting_info = True)
        for sheet in workbook.sheets():
            nrows = sheet.nrows
            #ncols = sheet.ncols
            if nrows==0:
                break #empty sheet
            row_list = []
            for row in range(sheet.nrows):
                col_list = []
                for col in range(sheet.ncols):
                    col_list.append(sheet.cell(row, col))
                row_list.append(col_list)
            filtered_merges = []
            for (pos_a,pos_b,pos_c,pos_d) in sheet.merged_cells:
                filtered_merges.append((pos_c,pos_a,pos_d-1,pos_b-1))
            sheet_list[sheet.name] = {'rows': row_list, 'merged_cells': filtered_merges}
        return sheet_list

    @staticmethod
    def load_file_xlsx(filename):
        """
        Open file using openpyxl for newer excel format
        """
        sheet_list={}
        workbook = openpyxl.load_workbook(filename)
        for sheet in workbook.worksheets:
            if sheet.max_row == 1:
                break
            bounds_list = []
            for i in sheet.merged_cells.ranges:
                bounds_list.append(i.bounds)
            row_list = []
            for row in sheet.iter_rows():
                col_list = []
                for cell in row:
                    mcell = {}
                    if cell.data_type=='n':
                        if cell.value:
                            mcell = SCell(2, cell.value)
                        else:
                            mcell = SCell(0, None)
                    elif cell.data_type=='s':
                        mcell = SCell(1, cell.value)
                    elif cell.data_type=='b':
                        mcell = SCell(4, cell.value)
                    elif cell.data_type=='e':
                        mcell = SCell(5, cell.value)
                    else:
                        mcell = SCell(0, None)
                    col_list.append(mcell)
                row_list.append(col_list)
            filtered_merges = []
            for (pos_a,pos_b,pos_c,pos_d) in bounds_list:
                filtered_merges.append((pos_a-1,pos_b-1,pos_c-1,pos_d-1))
            # above line changes filtered_merges to diffrent format
            print(filtered_merges)
            sheet_list[sheet.title] = {'rows': row_list, 'merged_cells': filtered_merges}
        return sheet_list

    @staticmethod
    def load_file_excel(filename):
        """
        Autodetermine which function to load based on extension
        """
        if filename.endswith('.xlsx'):
            return BoreholeDataSheets.load_file_xlsx(filename)
        return BoreholeDataSheets.load_file_xls(filename)

    @staticmethod
    def load_file_txt(filename):
        """
        Load multiple sheets from same location but have seperate filename
        Name Should be listed and seperated by new line character
        """
        output = {}
        with open(filename) as file:
            for line in file:
                filepath = Path(filename).parent / line.strip()
                result = BoreholeDataSheets.load_file_excel(str(filepath.resolve()))
                #enable excel file only no recursive
                for key_name in result:
                    generator = TempNameGenerator(key_name)
                    while True:#use generator to generate new name if duplicated
                        new_key_name = generator.next()
                        if new_key_name not in output.keys():
                            output[new_key_name] = result[key_name]
                            break
        return output

    @staticmethod
    def load_file(filename):
        """
        - Auto determine type from extension
        filename can be one of these three types
        .xlsx   -> newer excel format
        .xls    -> older excel format
        .txt    -> Contains multiple sheets from same location
        """
        if filename.endswith('.txt'):
            return BoreholeDataSheets.load_file_txt(filename)
        return BoreholeDataSheets.load_file_excel(filename)

if __name__ == "__main__":
    import doctest
    doctest.testmod()
